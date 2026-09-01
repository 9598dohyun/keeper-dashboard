"""
키퍼 주문정산통합데이터 엑셀 ↔ 에어테이블 대조

결제 여부의 진짜 소스는 이 엑셀이다.
엑셀에 없는 건은 에어테이블 [콜]최종 결과가 '결제 완료'여도 결제로 세지 않는다.

에어테이블 원본은 수정하지 않는다. 대조 결과만 data/결제대조.json 으로 떨어뜨리고,
compute-and-push 가 그 파일을 결제수 소스로 읽는다.

엑셀 구조 (첫 시트 '1. 주문건' 만 사용, '2. 추가 비용청구' 는 보지 않는다):
    주문일시 / 매장명 / 주문유입채널 / 주문금액 / 취소금액 / 주문상태 / 고객명 / 휴대폰번호 ...
파일에는 전체 기간이 담겨 있어(1년 반 누적) 기준일 건만 걸러서 쓴다.

사용법:
    python3 reconcile.py <주문정산통합데이터.xlsx> [--date YYYY-MM-DD] [--dry-run]

--date 생략 시 파일명의 날짜(…_20260826.xlsx)를 쓰고, 없으면 엑셀 내 최신 주문일을 쓴다.
"""

import argparse
import json
import os
import re
import unicodedata
import urllib.parse
import urllib.request
from collections import Counter
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

KST = timezone(timedelta(hours=9))
BASE_DIR = Path(__file__).resolve().parents[2]
OUT_PATH = BASE_DIR / "data" / "결제대조.json"
# 누적 결제 원장 — 엑셀을 받을 때마다 여기에 쌓인다.
# 하루치 파일만 와도 과거분이 남아 있어야 진단(유입 코호트)이 엑셀 기준으로 계산된다.
LEDGER_PATH = BASE_DIR / "data" / "결제원장.json"

INBOUND_TABLE = "tbljFHOl4PzAWmb1f"
SKB_TABLE = "tblb5APohbhFixfHB"

# 엑셀 컬럼 이름 후보 (매달 조금씩 바뀔 수 있어 후보를 둔다)
PHONE_HEADERS = ["휴대폰번호", "연락처", "휴대폰", "전화번호", "고객연락처"]
NAME_HEADERS = ["고객명", "이름", "성명", "가입자명"]
DATE_HEADERS = ["주문일시", "결제일", "결제일시", "카드 결제/취소일시"]
STORE_HEADERS = ["매장명", "상호명", "사업장명"]
STATUS_HEADERS = ["주문상태"]
CANCEL_HEADERS = ["취소금액"]
CHANNEL_HEADERS = ["주문유입채널"]
AMOUNT_HEADERS = ["주문금액"]
ORDER_NO_HEADERS = ["주문번호"]

# 주문상태에 이 말이 들어가면 취소 건 — 결제로 세지 않는다
CANCELLED_MARKERS = ("취소",)


def norm_header(s):
    """헤더 비교용 정규화 — 공백·괄호·구분기호 제거"""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    return re.sub(r"[\s\(\)\[\]/_.-]", "", s)


def phone_key(v):
    """
    전화번호 뒷 8자리 = 리드 대조 키.
    한화비전 중복 판정 기준과 같다. 하이픈·공백·국가번호가 섞여 오므로 숫자만 남긴다.
    """
    if v is None:
        return None
    digits = re.sub(r"\D", "", str(v))
    if len(digits) < 8:
        return None
    return digits[-8:]


def mask_name(n):
    """콘솔 표시용 이름 마스킹"""
    if not n:
        return "(이름없음)"
    n = str(n).strip()
    if len(n) <= 1:
        return n
    if len(n) == 2:
        return n[0] + "*"
    return n[0] + "*" * (len(n) - 2) + n[-1]


def find_col(headers, candidates):
    """후보 이름 중 먼저 맞는 컬럼 인덱스. 완전일치 우선, 없으면 부분일치"""
    normed = [norm_header(h) for h in headers]
    for cand in candidates:
        c = norm_header(cand)
        for i, h in enumerate(normed):
            if h == c:
                return i
    for cand in candidates:
        c = norm_header(cand)
        for i, h in enumerate(normed):
            if c and c in h:
                return i
    return None


def parse_date(v):
    """엑셀 셀 → 날짜(YYYY-MM-DD). '2026.08.26 18:38:30' 형식 포함. 판독 불가면 None"""
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    m = re.search(r"(\d{4})[-./년\s]*(\d{1,2})[-./월\s]*(\d{1,2})", s)
    if m:
        y, mo, d = (int(x) for x in m.groups())
        try:
            return datetime(y, mo, d).date().isoformat()
        except ValueError:
            return None
    return None


def date_from_filename(path):
    """파일명 안의 8자리 날짜(…_20260826.xlsx) → YYYY-MM-DD"""
    m = re.search(r"(20\d{2})(\d{2})(\d{2})", Path(path).stem)
    if not m:
        return None
    y, mo, d = (int(x) for x in m.groups())
    try:
        return datetime(y, mo, d).date().isoformat()
    except ValueError:
        return None


def read_excel(path):
    """첫 시트('1. 주문건')에서 주문 목록을 읽는다. 두 번째 시트는 보지 않는다."""
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet = wb.sheetnames[0]
    ws = wb[sheet]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    header_idx = None
    cols = None
    for i, row in enumerate(rows[:30]):
        p = find_col(row, PHONE_HEADERS)
        if p is not None:
            header_idx = i
            cols = {
                "phone": p,
                "name": find_col(row, NAME_HEADERS),
                "date": find_col(row, DATE_HEADERS),
                "store": find_col(row, STORE_HEADERS),
                "status": find_col(row, STATUS_HEADERS),
                "cancel": find_col(row, CANCEL_HEADERS),
                "channel": find_col(row, CHANNEL_HEADERS),
                "amount": find_col(row, AMOUNT_HEADERS),
                "order_no": find_col(row, ORDER_NO_HEADERS),
            }
            break

    if header_idx is None:
        raise SystemExit(
            f"엑셀 '{sheet}' 시트에서 휴대폰번호 컬럼을 못 찾았습니다.\n"
            f"  1행: {rows[0][:10] if rows else '(빈 파일)'}\n"
            f"  찾는 이름: {', '.join(PHONE_HEADERS)}\n"
            "  컬럼명이 바뀌었으면 이 스크립트의 PHONE_HEADERS에 추가해 주세요."
        )

    def cell(row, key):
        i = cols[key]
        if i is None or i >= len(row):
            return None
        return row[i]

    items = []
    skipped = 0
    for row in rows[header_idx + 1 :]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        key = phone_key(cell(row, "phone"))
        if key is None:
            skipped += 1
            continue
        status = cell(row, "status")
        name = cell(row, "name")
        store = cell(row, "store")
        channel = cell(row, "channel")
        order_no = cell(row, "order_no")
        items.append(
            {
                "주문번호": str(order_no).strip() if order_no else None,
                "키": key,
                "고객명": str(name).strip() if name else None,
                "매장명": str(store).strip() if store else None,
                "결제일": parse_date(cell(row, "date")),
                "주문상태": str(status).strip() if status else "",
                "취소금액": cell(row, "cancel"),
                "채널": str(channel).strip() if channel else "",
                "금액": cell(row, "amount"),
            }
        )

    header_names = {k: (rows[header_idx][v] if v is not None else None) for k, v in cols.items()}
    return items, skipped, header_names, sheet


def is_cancelled(it):
    """
    취소 건 판정 — 주문상태에 '취소'가 들어가거나 취소금액이 찍힌 건.

    취소금액은 음수로 들어온다(-1188000). 0이 아니면 취소로 본다.
    """
    if any(m in it["주문상태"] for m in CANCELLED_MARKERS):
        return True
    c = it["취소금액"]
    if c is None:
        return False
    try:
        return float(str(c).replace(",", "")) != 0
    except (TypeError, ValueError):
        return False


def dedupe_orders(items):
    """
    주문번호 1건당 1행으로 합친다.

    엑셀은 결제·취소를 각각 한 행으로 쌓아서, 취소된 주문은 같은 주문번호가 두 번 나온다
    (원 결제 행 + 취소 행). 그대로 세면 취소 건이 두 배가 되므로 주문번호로 접는다.
    취소 행이 하나라도 있으면 그 주문은 취소로 판정한다.
    """
    by_no = {}
    order = []
    for it in items:
        no = it["주문번호"]
        if not no:
            # 주문번호가 없으면 접을 근거가 없어 그대로 둔다
            order.append(it)
            continue
        if no not in by_no:
            by_no[no] = dict(it)
            order.append(by_no[no])
        elif is_cancelled(it):
            # 취소 행의 상태·금액을 살린다
            by_no[no]["주문상태"] = it["주문상태"]
            by_no[no]["취소금액"] = it["취소금액"]
    return order


def is_paid_result(최종결과):
    """
    최종결과가 결제 완료인지. 테이블마다 문구가 달라 접두어로 본다
    (인바운드 '결제 완료 (영원)' / SKB '결제 완료').
    """
    return str(최종결과 or "").startswith("결제 완료")


def pick_lead(hit):
    """
    주문 1건에 붙은 리드 여러 개 중 결제로 인정할 **하나**를 고른다.

    같은 고객이 여러 번 문의하면 같은 전화번호로 리드가 중복 생성된다. 붙은 리드를
    모두 결제로 세면 주문 1건이 리드 수만큼 계상된다
    (2026-09-01: 주문 24건 → 리드 32개. 인바운드 15→21, SKB 9→11로 부풀었다).

    선정 기준 (2026-09-02 확정):
      1) 최종결과가 '결제 완료'인 리드 우선
      2) 그중 유입시간이 가장 늦은 것 (= 마지막 유입 건)
      3) 결제 완료가 없으면 전체에서 유입시간이 가장 늦은 것

    테이블이 갈리는 경우(인바운드·SKB 양쪽에 같은 번호)는 **SKB의 결제 완료로 귀속**한다.
    """
    결제완료 = [r for r in hit if is_paid_result(r["최종결과"])]
    후보 = 결제완료 or hit
    skb_결제완료 = [r for r in 결제완료 if r["테이블"] == "SKB"]
    if skb_결제완료:
        후보 = skb_결제완료
    # 유입시간 내림차순 — 값이 없으면 빈 문자열이라 가장 뒤로 밀린다
    return max(후보, key=lambda r: r.get("유입시간") or "")


def load_ledger():
    """누적 결제 원장. 없으면 빈 원장"""
    if not LEDGER_PATH.exists():
        return {"주문": {}, "갱신이력": []}
    return json.loads(LEDGER_PATH.read_text(encoding="utf-8"))


def update_ledger(ledger, orders, index, 엑셀파일):
    """
    원장에 이번 엑셀의 주문을 반영한다.

    같은 주문번호는 덮어쓴다 — 어제 결제였다가 오늘 취소된 건이 취소로 갱신돼야 한다.
    개인정보는 남기지 않고 결제일·취소여부·매칭된 레코드 ID만 저장한다.
    """
    주문 = ledger.setdefault("주문", {})
    added = updated = 0
    for it in orders:
        no = it["주문번호"]
        if not no:
            continue
        hit = index.get(it["키"], [])
        # 주문 1건 = 결제 1건. 중복 리드가 붙어도 대표 1개만 원장에 남긴다
        대표 = [pick_lead(hit)] if hit else []
        rec = {
            "결제일": it["결제일"],
            "취소": is_cancelled(it),
            "채널": it["채널"],
            "인바운드ID": sorted({r["id"] for r in 대표 if r["테이블"] == "인바운드"}),
            "skbID": sorted({r["id"] for r in 대표 if r["테이블"] == "SKB"}),
            "매칭": bool(hit),
        }
        if no in 주문:
            if 주문[no] != rec:
                주문[no] = rec
                updated += 1
        else:
            주문[no] = rec
            added += 1
    ledger["갱신이력"] = (ledger.get("갱신이력", []) + [
        {
            "시각": datetime.now(KST).isoformat(),
            "엑셀파일": 엑셀파일,
            "신규": added,
            "갱신": updated,
        }
    ])[-30:]
    return added, updated


def ledger_payment_ids(ledger):
    """원장 → 결제로 인정되는 레코드 ID 집합 (취소 제외)"""
    inb, skb = set(), set()
    for rec in ledger.get("주문", {}).values():
        if rec.get("취소"):
            continue
        inb.update(rec.get("인바운드ID", []))
        skb.update(rec.get("skbID", []))
    return inb, skb


def airtable_fetch(base, token, table, fields, progress=None):
    """테이블 전체 조회 (페이징). 인바운드는 1만 건대라 수십 초 걸린다"""
    out = []
    offset = None
    page = 0
    while True:
        q = [("pageSize", "100")] + [("fields[]", f) for f in fields]
        if offset:
            q.append(("offset", offset))
        url = f"https://api.airtable.com/v0/{base}/{table}?" + urllib.parse.urlencode(q)
        req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
        with urllib.request.urlopen(req, timeout=90) as r:
            d = json.load(r)
        out += d.get("records", [])
        offset = d.get("offset")
        page += 1
        if progress and page % 20 == 0:
            print(f"    {progress}: {len(out)}건...", flush=True)
        if not offset:
            return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("excel", help="키퍼 주문정산통합데이터 엑셀 경로")
    ap.add_argument("--date", help="기준일 YYYY-MM-DD (생략 시 파일명 날짜 → 엑셀 최신 주문일)")
    ap.add_argument("--dry-run", action="store_true", help="파일로 저장하지 않고 결과만 출력")
    ap.add_argument(
        "--all",
        action="store_true",
        help=(
            "엑셀에 담긴 전체 기간을 원장에 반영한다. "
            "전체 기간이 담긴 파일을 처음 넣을 때(소급 적용) 사용."
        ),
    )
    ap.add_argument(
        "--exclude",
        default="",
        help=(
            "결제에서 뺄 건. 주문번호 또는 매장명을 쉼표로 구분. "
            "엑셀만 봐서는 알 수 없는 건(예: 이전 결제분을 뒤늦게 관리자등록한 건)을 손으로 뺄 때 쓴다."
        ),
    )
    args = ap.parse_args()

    token = os.environ.get("AIRTABLE_TOKEN")
    base = os.environ.get("AIRTABLE_SKB_BASE_ID") or os.environ.get("AIRTABLE_BASE_ID")
    if not token or not base:
        raise SystemExit("AIRTABLE_TOKEN / AIRTABLE_SKB_BASE_ID 환경변수가 필요합니다.")

    all_items, skipped, hdr, sheet = read_excel(args.excel)
    if not all_items:
        raise SystemExit("엑셀에서 읽은 주문이 0건입니다. 파일을 확인해 주세요.")

    print(f"엑셀: {Path(args.excel).name}  (시트 '{sheet}')")
    print(
        f"  컬럼 — 번호={hdr['phone']!r} 고객명={hdr['name']!r} 일시={hdr['date']!r} "
        f"상태={hdr['status']!r} 취소={hdr['cancel']!r}"
    )
    print(f"  전체 주문 {len(all_items)}건" + (f" (번호 없어 제외 {skipped}행)" if skipped else ""))

    dates = sorted({i["결제일"] for i in all_items if i["결제일"]})
    if dates:
        print(f"  주문일 범위: {dates[0]} ~ {dates[-1]} ({len(dates)}일치 누적)")

    # 기준일 결정.
    # 파일명 날짜(…_20260827)는 '내려받은 날'이라 주문일과 다를 수 있다
    # (전날 마감분을 다음날 아침에 받는 경우). 파일명 날짜에 주문이 없으면
    # 엑셀에 실제로 담긴 최신 주문일로 넘어간다.
    from_name = date_from_filename(args.excel)
    if args.date:
        기준일, 출처 = args.date, "지정"
    elif from_name and from_name in dates:
        기준일, 출처 = from_name, "파일명"
    elif dates:
        기준일, 출처 = dates[-1], "엑셀 최신 주문일"
        if from_name:
            print(f"  참고: 파일명 날짜 {from_name} 주문이 없어 엑셀 최신 주문일을 씁니다.")
    else:
        raise SystemExit("기준일을 정할 수 없습니다. --date 로 지정해 주세요.")
    print(f"  기준일: {기준일} ({출처} 기준)")

    # 기준일 건만 사용 — 파일에는 전체 기간이 담겨 있다
    day = [i for i in all_items if i["결제일"] == 기준일]
    if not day and not args.all:
        raise SystemExit(
            f"기준일 {기준일} 주문이 엑셀에 없습니다.\n"
            f"  엑셀에 담긴 최근 날짜: {', '.join(dates[-5:]) if dates else '(없음)'}\n"
            "  --date 로 다른 날짜를 지정하거나 파일을 확인해 주세요."
        )

    # --all 이면 파일에 담긴 전체 기간을 원장에 넣는다 (소급 적용)
    if args.all:
        day = all_items
        print(f"\n--all: 전체 기간 {len(dates)}일치를 원장에 반영합니다.")
    rows_raw = len(day)
    day = dedupe_orders(day)

    # 손으로 빼는 건 — 엑셀 데이터만으로는 판별할 수 없는 예외를 처리한다.
    # (예: 이전에 결제된 건을 뒤늦게 '관리자등록'으로 넣어 오늘 주문처럼 보이는 경우)
    제외키 = {x.strip() for x in args.exclude.split(",") if x.strip()}
    제외됨 = []
    if 제외키:
        keep = []
        for it in day:
            if it["주문번호"] in 제외키 or (it["매장명"] and it["매장명"] in 제외키):
                제외됨.append(it)
            else:
                keep.append(it)
        day = keep
        미발견 = 제외키 - {i["주문번호"] for i in 제외됨} - {i["매장명"] for i in 제외됨 if i["매장명"]}
        if 미발견:
            raise SystemExit(
                f"--exclude 로 지정한 값을 기준일 주문에서 못 찾았습니다: {', '.join(sorted(미발견))}\n"
                "  주문번호나 매장명이 정확한지 확인해 주세요."
            )
    유효 = [i for i in day if not is_cancelled(i)]
    취소 = [i for i in day if is_cancelled(i)]
    dedup_note = (
        f" (엑셀 {rows_raw}행 → 주문 {len(day) + len(제외됨)}건)"
        if rows_raw != len(day) + len(제외됨)
        else ""
    )
    print(f"\n{기준일} 주문 {len(day)}건{dedup_note} → 유효 결제 {len(유효)}건 / 취소 {len(취소)}건")
    if 취소:
        상태들 = sorted(set(i["주문상태"] for i in 취소 if i["주문상태"]))
        print(f"  취소 내역: {', '.join(상태들)}")
    if 제외됨:
        print(f"  손으로 제외 {len(제외됨)}건: " + ", ".join(
            f"{i['매장명'] or i['주문번호']}" for i in 제외됨
        ))
    ch = Counter(i["채널"] or "(없음)" for i in 유효)
    print(f"  유입채널: {', '.join(f'{k} {v}' for k, v in ch.most_common())}")

    print("\n에어테이블 조회 중...")
    inbound = airtable_fetch(
        base,
        token,
        INBOUND_TABLE,
        ["연락처", "고객명", "[콜]최종 결과", "유입시간"],
        "인바운드",
    )
    skb = airtable_fetch(
        base, token, SKB_TABLE, ["연락처", "이름", "[콜]최종 결과", "유입시간"], "SKB"
    )
    print(f"  인바운드 {len(inbound)}건 / SKB {len(skb)}건")

    index = {}
    for label, recs, name_field in (("인바운드", inbound, "고객명"), ("SKB", skb, "이름")):
        for r in recs:
            f = r.get("fields", {})
            k = phone_key(f.get("연락처"))
            if k is None:
                continue
            index.setdefault(k, []).append(
                {
                    "테이블": label,
                    "id": r["id"],
                    "고객명": f.get(name_field),
                    "최종결과": f.get("[콜]최종 결과"),
                    "유입시간": f.get("유입시간") or "",
                }
            )

    매칭, 미매칭 = [], []
    중복리드 = []
    for it in 유효:
        hit = index.get(it["키"])
        if hit:
            # 주문 1건 = 결제 1건. 중복 리드가 붙어도 대표 1개만 결제로 센다
            대표 = pick_lead(hit)
            if len(hit) > 1:
                중복리드.append((it, hit, 대표))
            매칭.append({**it, "리드": [대표]})
        else:
            미매칭.append(it)

    by_table = Counter(r["테이블"] for m in 매칭 for r in m["리드"])

    print("\n=== 대조 결과 ===")
    print(f"{기준일} 유효 결제 {len(유효)}건")
    print(
        f"  에어테이블 매칭   {len(매칭)}건"
        + (f"  ({', '.join(f'{k} {v}' for k, v in sorted(by_table.items()))})" if by_table else "")
    )
    print(f"  에어테이블 미매칭 {len(미매칭)}건" + ("  <== 리드 없이 결제된 건" if 미매칭 else ""))

    if 중복리드:
        print(f"\n[중복 리드 — 주문 1건에 리드 여러 개, 대표 1개만 결제로 셈] {len(중복리드)}건")
        for it, hit, 대표 in 중복리드:
            others = Counter(r["테이블"] for r in hit)
            print(
                f"  {it['매장명'] or it['주문번호']}  리드 {len(hit)}개({dict(others)})"
                f" → {대표['테이블']} {대표['최종결과'] or '(결과없음)'}"
            )

    ch_매칭 = Counter(m["채널"] or "(없음)" for m in 매칭)
    print("\n[유입채널별 결제]")
    for k, v in ch.most_common():
        print(f"  {k:12} 유효 {v}건  (매칭 {ch_매칭.get(k, 0)}건)")

    if 미매칭:
        print("\n[미매칭 — 엑셀에 있으나 에어테이블에 리드 없음]")
        for m in 미매칭[:20]:
            print(
                f"  ...{m['키'][-4:]}  {mask_name(m['고객명'])}  "
                f"{m['매장명'] or ''}  채널={m['채널'] or '?'}"
            )
        if len(미매칭) > 20:
            print(f"  … 외 {len(미매칭) - 20}건")

    if 취소:
        print("\n[취소 — 결제로 세지 않음]")
        for c in 취소[:10]:
            print(f"  ...{c['키'][-4:]}  {mask_name(c['고객명'])}  {c['매장명'] or ''}  {c['주문상태']}")
        if len(취소) > 10:
            print(f"  … 외 {len(취소) - 10}건")

    # 누적 원장 갱신 — 진단(유입 코호트)이 과거 결제까지 엑셀 기준으로 보게 한다
    ledger = load_ledger()
    added, updated = update_ledger(ledger, day, index, Path(args.excel).name)
    inb_ids, skb_ids = ledger_payment_ids(ledger)
    if not args.dry_run:
        LEDGER_PATH.parent.mkdir(parents=True, exist_ok=True)
        LEDGER_PATH.write_text(
            json.dumps(ledger, ensure_ascii=False, indent=1), encoding="utf-8"
        )
    print(
        f"\n원장: 주문 {len(ledger['주문'])}건 (신규 {added} / 갱신 {updated})"
        f" → 결제 인정 인바운드 {len(inb_ids)} · SKB {len(skb_ids)}"
    )

    payload = {
        "기준일": 기준일,
        "생성시각": datetime.now(KST).isoformat(),
        "엑셀파일": Path(args.excel).name,
        "결제_전체": len(유효),
        "결제_매칭": len(매칭),
        "취소_건수": len(취소),
        "미매칭_건수": len(미매칭),
        "수동제외_건수": len(제외됨),
        "수동제외": sorted(
            f"{i['주문번호']}({i['매장명']})" if i["매장명"] else str(i["주문번호"])
            for i in 제외됨
        ),
        # 유입채널별 결제건 — 엑셀 `주문유입채널` 기준.
        # 전체는 유효 결제, 매칭은 그중 리드가 붙은 것(대시보드 결제수와 같은 기준)
        "채널별_결제": dict(ch.most_common()),
        "채널별_결제_매칭": dict(Counter(m["채널"] or "(없음)" for m in 매칭).most_common()),
        "채널별_취소": dict(Counter(c["채널"] or "(없음)" for c in 취소).most_common()),
        # 개인정보(이름·연락처)는 파일에 남기지 않는다 — 레코드 ID와 건수만
        "결제ID_인바운드": sorted(
            {r["id"] for m in 매칭 for r in m["리드"] if r["테이블"] == "인바운드"}
        ),
        "결제ID_SKB": sorted({r["id"] for m in 매칭 for r in m["리드"] if r["테이블"] == "SKB"}),
        # 원장 누적분 — 진단 화면이 과거 코호트를 엑셀 기준으로 셀 때 쓴다
        "원장_주문수": len(ledger["주문"]),
        "원장_결제ID_인바운드": sorted(inb_ids),
        "원장_결제ID_SKB": sorted(skb_ids),
    }

    if args.dry_run:
        print("\n--dry-run: 파일 저장 생략")
        return

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"\n저장: {OUT_PATH.relative_to(BASE_DIR)}")
    print("다음: npx tsx scripts/fetch-airtable.ts && npx tsx scripts/compute-and-push.ts")


if __name__ == "__main__":
    main()
